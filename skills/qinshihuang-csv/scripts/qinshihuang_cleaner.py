#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
qinshihuang-csv — 电商CSV大一统清洗方案
========================================
两步走：Step 1 编码检测与转换 → Step 2 字段级脏数据清洗
"""

import os
import re
import sys
import codecs
import logging
import argparse
import datetime
from pathlib import Path

import pandas as pd
from charset_normalizer import from_path as cn_from_path
import chardet

# ── 模块级日志 ──
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    datefmt="%H:%M:%S",
)
log = logging.getLogger("qinshihuang")

# ── 特殊字符定义 ──
_CURRENCY_SYMBOLS = "￥$¥€£₩₽₪₫₱₹₴₸₺₼₿₠₡₢₣₤₥₦₧₨₰₲₳₵₶₷₻₾"
_STRIP_CHARS = f",;\t\n\r {_CURRENCY_SYMBOLS}"

# ── 广度回退编码列表 ──
FALLBACK_ENCODINGS = [
    "utf-8-sig", "utf-8", "gb18030", "gbk", "gb2312",
    "big5", "big5hkscs",
    "shift_jis", "euc-jp", "iso-2022-jp",
    "euc-kr", "iso-2022-kr",
    "utf-16", "utf-16-le", "utf-16-be",
    "utf-32", "utf-32-le", "utf-32-be",
    "cp1252", "cp1250", "cp1251", "cp1253", "cp1254", "cp1255", "cp1256", "cp1257", "cp1258",
    "iso-8859-1", "iso-8859-2", "iso-8859-3", "iso-8859-4", "iso-8859-5",
    "iso-8859-6", "iso-8859-7", "iso-8859-8", "iso-8859-9", "iso-8859-10",
    "iso-8859-13", "iso-8859-14", "iso-8859-15", "iso-8859-16",
    "latin1", "mac_roman", "mac_cyrillic", "mac_greek", "mac_iceland", "mac_turkish",
    "cp437", "cp850", "cp852", "cp855", "cp857", "cp858", "cp860", "cp861",
    "cp862", "cp863", "cp864", "cp865", "cp866", "cp869",
    "koi8-r", "koi8-u",
    "hz", "iso-2022-cn", "euc-tw",
]

# ── 可选依赖：CleverCSV 方言检测 ──
try:
    import clevercsv
    HAS_CLEVERCSV = True
except ImportError:
    HAS_CLEVERCSV = False


# ═══════════════════════════════════════════════════
#  Step 1: 编码检测与转换
# ═══════════════════════════════════════════════════

def detect_bom(file_path: str) -> str | None:
    """检测 BOM 头，返回编码名称或 None。"""
    with open(file_path, "rb") as f:
        raw = f.read(4)

    if raw.startswith(codecs.BOM_UTF8):
        return "utf-8-sig"
    if raw[:4] == b"\xff\xfe\x00\x00":
        return "utf-32-le"
    if raw[:4] == b"\x00\x00\xfe\xff":
        return "utf-32-be"
    if raw[:2] == codecs.BOM_UTF16_LE:
        return "utf-16-le"
    if raw[:2] == codecs.BOM_UTF16_BE:
        return "utf-16-be"
    return None


def detect_encoding_charset_normalizer(file_path: str, sample_size: int = 100_000) -> str | None:
    """使用 charset-normalizer 检测编码。"""
    try:
        # 不同版本 API 有差异：从文件读取样本再传 bytes
        with open(file_path, "rb") as f:
            raw_sample = f.read(sample_size)
        result = cn_from_path(file_path).best()
        if result is None:
            # 回退：用 bytes 检测
            from charset_normalizer import from_bytes
            result = from_bytes(raw_sample).best()
        if result and result.encoding and result.confidence is not None:
            if result.confidence >= 0.7:
                enc = result.encoding.lower()
                # GB2312/GBK → GB18030（超集兼容）
                if enc in ("gb2312", "gbk"):
                    return "gb18030"
                return enc
    except Exception:
        pass
    return None


def detect_encoding_chardet(file_path: str, sample_size: int = 100_000) -> str | None:
    """使用 chardet 兜底检测。

    对中文内容（language='zh'）降置信度门槛到 0.2，
    因为电商CSV常有少量坏字节但仍可采取 GB18030 解码。"""
    try:
        with open(file_path, "rb") as f:
            raw = f.read(sample_size)
        result = chardet.detect(raw)
        enc = result.get("encoding")
        confidence = result.get("confidence", 0)
        lang = result.get("language", "")
        if enc:
            # 中文内容降低门槛
            threshold = 0.2 if lang == "zh" else 0.5
            if confidence >= threshold:
                enc = enc.lower()
                if enc in ("gb2312", "gbk"):
                    return "gb18030"
                return enc
    except Exception:
        pass
    return None


def try_fallback_encodings(file_path: str) -> str | None:
    """广度回退列表：依次尝试解码前两行。

    对中文编码（gb18030/gbk/big5等），如果严格模式失败，
    尝试 errors='replace' 模式并验证输出是否含 CJK 字符。"""
    with open(file_path, "rb") as f:
        raw_header = f.read(4096)

    # CJK 编码组：这些编码可尝试容错模式
    cjk_encodings = {"gb18030", "gbk", "gb2312", "big5", "big5hkscs", "euc-tw", "euc-jp", "shift_jis"}

    for enc in FALLBACK_ENCODINGS:
        try:
            raw_header.decode(enc)
            # 进一步确认：尝试读取文件的前几行
            with open(file_path, "r", encoding=enc) as f:
                f.readline()
                f.readline()
            return enc
        except (UnicodeDecodeError, LookupError, UnicodeError):
            # 对 CJK 编码尝试容错模式
            if enc in cjk_encodings:
                try:
                    decoded = raw_header.decode(enc, errors="replace")
                    # 检查是否有 CJK 字符（U+4E00-U+9FFF 基本汉字）
                    cjk_count = sum(1 for c in decoded if '\u4e00' <= c <= '\u9fff')
                    if cjk_count >= 5:  # 至少 5 个汉字
                        return enc
                except Exception:
                    continue
    return None


def detect_encoding(file_path: str) -> str:
    """四层编码检测：BOM → charset-normalizer → chardet → 回退列表。"""
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在：{file_path}")
    if path.stat().st_size == 0:
        raise ValueError(f"文件为空：{file_path}")

    # Layer 1: BOM
    bom_enc = detect_bom(file_path)
    if bom_enc:
        log.info(f"📌 BOM 头检测 → {bom_enc}")
        return bom_enc

    # Layer 2: charset-normalizer
    cn_enc = detect_encoding_charset_normalizer(file_path)
    if cn_enc:
        log.info(f"✅ charset-normalizer 检测 → {cn_enc}")
        return cn_enc

    # Layer 3: chardet
    cd_enc = detect_encoding_chardet(file_path)
    if cd_enc:
        log.info(f"✅ chardet 检测 → {cd_enc}")
        return cd_enc

    # Layer 4: 广度回退
    fb_enc = try_fallback_encodings(file_path)
    if fb_enc:
        log.info(f"✅ 回退列表匹配 → {fb_enc}")
        return fb_enc

    raise ValueError(f"无法识别文件的编码格式：{file_path}")


def read_csv_with_encoding(file_path: str, encoding: str) -> tuple[pd.DataFrame, str]:
    """用指定编码读取 CSV，所有行作为纯数据返回（不自动识别表头）。

    如果严格解码失败，自动回退到 errors='replace' 模式。"""
    log.info(f"📖 用 {encoding} 读取文件...")
    try:
        # 分隔符检测：CleverCSV > 简单启发式
        sep = ","
        if HAS_CLEVERCSV:
            try:
                dialect = clevercsv.detect_dialect(file_path, encoding=encoding)
                sep = dialect.delimiter
                log.info(f"  方言检测 (CleverCSV)：delimiter={repr(sep)}, "
                         f"quote={repr(dialect.quotechar)}, "
                         f"escape={repr(dialect.escapechar)}")
            except Exception:
                pass

        if not sep or sep == "\x00":
            sep = ","

        with open(file_path, "r", encoding=encoding) as f:
            first_line = f.readline()
        if sep == "," and "\t" in first_line and "," not in first_line[:200]:
            sep = "\t"
        log.info(f"  分隔符：{repr(sep)}")

        # 尝试严格模式，如果失败则用 errors='replace' 容错
        for errors_mode in ["strict", "replace"]:
            try:
                df = pd.read_csv(
                    file_path,
                    encoding=encoding,
                    encoding_errors=errors_mode,
                    sep=sep,
                    dtype=str,
                    keep_default_na=False,
                    na_filter=False,
                    header=None,
                    on_bad_lines="skip",
                )
                break
            except UnicodeDecodeError:
                if errors_mode == "replace":
                    log.warning(f"  ⚠️ 严格模式解码失败，改用 errors='replace' 容错")
                continue

        log.info(f"  读取完成：{df.shape[0]} 行 × {df.shape[1]} 列")
        return df, sep
    except Exception as e:
        raise RuntimeError(f"读取文件失败（{encoding}）：{e}")


# ═══════════════════════════════════════════════════
#  Step 2: 数据清洗
# ═══════════════════════════════════════════════════

def auto_detect_header(df: pd.DataFrame) -> int:
    """自动识别表头行。

    规则：从第0行开始扫描，找第一个所有单元格都是"干净文本"
    （不包含脏特殊字符模式）的行作为表头。
    如果找不到，返回 0（视第一行为表头）。
    """
    for row_idx in range(min(5, len(df))):
        row = df.iloc[row_idx]
        # 检查这一行是否"干净"——所有非空单元格都是普通文本
        all_clean = True
        for val in row:
            val_str = str(val).strip()
            if not val_str:
                continue
            # 表头不应该有: 首尾特殊字符、纯数字等
            # 简单判断：如果值以特殊字符开头或结尾，或者全是数字，不是好表头
            if re.match(r"^[\d\s,;\t￥$¥€£]+$", val_str):
                all_clean = False
                break
        if all_clean and row.notna().sum() >= 2:  # 至少2个非空列
            return row_idx
    return 0


def clean_cell_text(text: str) -> str:
    """清洗单个单元格：移除制表符/换行符 + 循环清除首尾特殊字符。"""
    if not isinstance(text, str):
        return str(text) if text is not None else ""
    if not text:
        return text

    # Step 1: 移除内部制表符、换行符、回车符
    text = text.replace("\t", "").replace("\n", "").replace("\r", "")

    # Step 2: 循环清除首尾特殊字符（因为可能复合出现）
    prev = None
    while prev != text:
        prev = text
        text = text.strip(_STRIP_CHARS)

    return text


def is_pure_digits_15plus(s: str) -> bool:
    """判断清洗后的字符串是否为 12+ 位连续数字（Excel 精度限制）。
    
    Excel 只能精确表示 15 位有效数字，超出的后几位会变成 0。
    从 12 位开始保护（含 12 位），覆盖订单号/身份证/流水号等常见场景。"""
    return bool(re.match(r"^\d{12,}$", s))


_DATE_PATTERNS = [
    r"^\d{4}-\d{1,2}-\d{1,2}$",           # 2025-10-19
    r"^\d{4}/\d{1,2}/\d{1,2}$",            # 2025/10/19
    r"^\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{2}$",         # 2025-10-19 14:25
    r"^\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{2}:\d{2}$",   # 2025-10-19 14:25:59
    r"^\d{4}/\d{1,2}/\d{1,2} \d{1,2}:\d{2}:\d{2}$",   # 2025/10/19 14:25:59
]


def is_date_string(s: str) -> bool:
    """判断字符串是否为常见日期格式。"""
    return any(re.match(p, s) for p in _DATE_PATTERNS)


def is_numeric_string(s: str) -> bool:
    """判断字符串是否为可解析的数值（含负数、小数）。"""
    if not s or s.strip() in ("", "-", "–"):
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False


def is_pure_dash(s: str) -> bool:
    """判断字符串是否仅为短横/减号（金额列中表示0）。"""
    return s.strip() in ("-", "–", "—")


def scan_dirty_patterns(df: pd.DataFrame, sample_rows: int = 15) -> tuple[dict, list, list, list, list]:
    """扫描前 N 行数据，确定每列的清洗策略、文本/日期/数值格式。

    返回：
        (
            {col_name: {"needs_clean": bool, "text_format": bool,
                         "date_format": bool, "numeric_format": bool}},
            [text_format_cols],
            [date_format_cols],
            [numeric_format_cols],
            [dash_clean_cols],  # 含短横需要替换为0的列
        )
    """
    sample = df.head(sample_rows).copy()

    patterns = {}
    for col in df.columns:
        col_data = sample[col].dropna().astype(str)
        needs_clean = False
        text_format = False
        date_format = False
        numeric_format = False
        has_decimal = False
        has_dash = False

        for val in col_data:
            if not val:
                continue
            cleaned = clean_cell_text(val)
            if cleaned != val:
                needs_clean = True

            # 文本格式检测（15+位纯数字，防科学计数法）
            if is_pure_digits_15plus(cleaned):
                text_format = True

            # 日期格式检测
            if is_date_string(cleaned) or is_date_string(val):
                date_format = True

            # 数值格式检测（含负数、小数）
            if is_numeric_string(cleaned):
                numeric_format = True
            # 含小数点（金额特征）
            if "." in cleaned and is_numeric_string(cleaned):
                has_decimal = True

            # 检测纯短横（金额列占位）
            if is_pure_dash(cleaned) or is_pure_dash(val):
                has_dash = True

        patterns[col] = {
            "needs_clean": needs_clean,
            "text_format": text_format,
            "date_format": date_format,
            "numeric_format": numeric_format,
            "has_decimal": has_decimal,
            "has_dash": has_dash,
        }

    text_format_cols = [c for c, p in patterns.items() if p["text_format"]]
    date_format_cols = [c for c, p in patterns.items() if p["date_format"] and not p["text_format"]]
    date_or_text_set = set(text_format_cols) | set(date_format_cols)
    # 数值格式：至少有一个值带小数点（金额特征），且不在文本/日期列中
    numeric_format_cols = [c for c, p in patterns.items()
                           if p.get("has_decimal") and c not in date_or_text_set]
    # 短横替换覆盖：数值列 + 有短横的非文本/非日期列
    dash_clean_cols = [c for c, p in patterns.items()
                       if p.get("has_dash") and c not in date_or_text_set]

    if text_format_cols:
        log.info(f"  🔢 文本格式列（防科学计数法）：{text_format_cols}")
    if date_format_cols:
        log.info(f"  📅 日期格式列：{date_format_cols}")
    if numeric_format_cols:
        log.info(f"  💰 数值格式列：{numeric_format_cols}")

    return patterns, text_format_cols, date_format_cols, numeric_format_cols, dash_clean_cols


def clean_dataframe(df: pd.DataFrame, patterns: dict, dash_clean_cols: list) -> pd.DataFrame:
    """全量清洗 DataFrame。

    对数值格式列中的纯短横/减号替换为 0。"""
    df_clean = df.copy()
    total_cells = 0
    cleaned_cells = 0
    dash_fixes = 0

    dash_set = set(dash_clean_cols)

    for col in df_clean.columns:
        for idx in range(len(df_clean)):
            orig = df_clean.at[idx, col]
            if not isinstance(orig, str) or not orig:
                continue
            total_cells += 1
            cleaned = clean_cell_text(orig)
            if cleaned != orig:
                df_clean.at[idx, col] = cleaned
                cleaned_cells += 1
            # 数值列：纯短横 → 0
            if col in dash_set and is_pure_dash(cleaned):
                df_clean.at[idx, col] = "0"
                dash_fixes += 1

    log.info(f"  🧹 清洗完成：{cleaned_cells}/{total_cells} 个单元格被清理"
             f"{f'，{dash_fixes} 个短横→0' if dash_fixes else ''}")
    return df_clean


# ═══════════════════════════════════════════════════
#  Step 3: 输出
# ═══════════════════════════════════════════════════

def make_output_path(original_path: str, suffix: str, output_dir: str = None) -> str:
    """生成输出文件路径。"""
    p = Path(original_path)
    stem = p.stem  # 不含扩展名
    if output_dir:
        out_dir = Path(output_dir)
    else:
        out_dir = p.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    return str(out_dir / f"{stem}{suffix}")


def write_csv(df: pd.DataFrame, output_path: str, encoding: str = "utf-8-sig",
              text_format_cols: list = None):
    """写入 CSV，用目标编码。

    对文本格式列（防科学计数法）使用前导单引号，
    Excel 打开时单引号不显示、单元格保持文本、不转科学计数法。"""
    df_out = df.copy()
    text_set = set(text_format_cols or [])

    if text_set:
        for col in df_out.columns:
            if col in text_set:
                df_out[col] = df_out[col].apply(
                    lambda v: f"'{v}" if v and str(v).strip() else v
                )

    df_out.to_csv(output_path, index=False, encoding=encoding)
    log.info(f"  💾 CSV 已写入：{output_path}")


def _parse_date_safe(s):
    """尝试将字符串解析为 datetime，失败返回原值。"""
    if not s or not isinstance(s, str):
        return s
    s = s.strip()
    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
    ]:
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s


def write_xlsx(df: pd.DataFrame, output_path: str,
               text_format_cols: list = None,
               date_format_cols: list = None,
               numeric_format_cols: list = None):
    """写入 XLSX，Excel 2016+ 兼容。

    三种列格式：
    - 文本格式 — 前导单引号 + @ 格式，防科学计数法
    - 日期格式 — 值转为 datetime 对象 + YYYY-MM-DD HH:MM:SS
    - 数值格式 — 值转为 float + #,##0.00
    """
    from openpyxl import Workbook
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active

    headers = list(df.columns)
    ws.append(headers)

    text_set = set(text_format_cols or [])
    date_set = set(date_format_cols or [])
    numeric_set = set(numeric_format_cols or [])

    text_col_indices = {i for i, h in enumerate(headers, 1) if h in text_set}
    date_col_indices = {i for i, h in enumerate(headers, 1) if h in date_set and h not in text_set}
    numeric_col_indices = {i for i, h in enumerate(headers, 1)
                           if h in numeric_set and h not in text_set and h not in date_set}

    # 逐行写入，按列类型转换值
    for _, row in df.iterrows():
        values = []
        for col_idx, (col_name, val) in enumerate(zip(headers, row), start=1):
            if col_idx in date_col_indices:
                values.append(_parse_date_safe(val))
            elif col_idx in numeric_col_indices:
                try:
                    v = float(val) if val and str(val).strip() not in ("", "-", "–") else 0.0
                    values.append(v)
                except (ValueError, TypeError):
                    values.append(0.0)
            elif col_idx in text_col_indices:
                values.append(val)
            else:
                values.append(val)
        ws.append(values)

    # 设置单元格格式
    for col_idx in text_col_indices:
        col_letter = get_column_letter(col_idx)
        for row_num in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row_num}"].number_format = '@'

    for col_idx in date_col_indices:
        col_letter = get_column_letter(col_idx)
        for row_num in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row_num}"].number_format = 'YYYY-MM-DD HH:MM:SS'

    for col_idx in numeric_col_indices:
        col_letter = get_column_letter(col_idx)
        for row_num in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row_num}"].number_format = '#,##0.00'

    # 列宽自适应
    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for row_num in range(2, min(ws.max_row + 1, 102)):
            cell_val = ws[f"{col_letter}{row_num}"].value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)

    wb.save(output_path)
    log.info(f"  💾 XLSX 已写入：{output_path}")


# ═══════════════════════════════════════════════════
#  主流程
# ═══════════════════════════════════════════════════

def process_csv(
    file_path: str,
    target_encoding: str = "utf-8-sig",
    to_xlsx: bool = False,
    output_dir: str = None,
    interactive: bool = False,
    desktop: str = None,
) -> dict:
    """处理单个 CSV 文件。

    返回结果字典供 Hermes Agent 汇报。
    """
    log.info(f"\n{'='*60}")
    log.info(f"📄 处理文件：{file_path}")
    log.info(f"{'='*60}")

    result = {
        "file": file_path,
        "status": "ok",
        "encoding": None,
        "encoding_method": None,
        "shape": None,
        "cleaned_cells": 0,
        "text_format_cols": [],
        "output_csv": None,
        "output_xlsx": None,
    }

    # ── Step 1: 编码检测 ──
    detected_enc = detect_encoding(file_path)
    result["encoding"] = detected_enc
    log.info(f"🎯 最终编码：{detected_enc}")

    # ── 读取 ──
    df, sep = read_csv_with_encoding(file_path, detected_enc)
    result["shape"] = f"{df.shape[0]}行×{df.shape[1]}列"

    if df.empty:
        log.warning("  ⚠️ 文件为空，跳过清洗")
        result["status"] = "empty"
        return result

    # ── 目标编码判断 ──
    # 对比标准化后的编码名
    norm_detected = detected_enc.lower().replace("-", "").replace("_", "")
    norm_target = target_encoding.lower().replace("-", "").replace("_", "")
    if norm_detected == norm_target:
        log.info(f"  ℹ️ 文件编码已是目标编码 ({target_encoding})")
        if interactive:
            return {"status": "confirm_needed", "encoding": detected_enc,
                    "message": f"文件编码已经是 {target_encoding}，是否继续执行清洗？"}
        else:
            log.info("  非交互模式，自动继续...")

    # ── Step 2: 自动识别表头并重构 DataFrame ──
    header_row = auto_detect_header(df)
    log.info(f"  📋 表头所在行：第 {header_row + 1} 行")

    # 用表头行设置列名
    headers = [str(v).strip() for v in df.iloc[header_row].tolist()]
    df.columns = headers
    log.info(f"  表头内容：{headers}")

    # 移除表头行，只留数据行
    df_data = df.iloc[header_row + 1:].reset_index(drop=True)

    # ── 扫描脏数据特征 ──
    patterns, text_format_cols, date_format_cols, numeric_format_cols, dash_clean_cols = \
        scan_dirty_patterns(df_data)
    result["text_format_cols"] = text_format_cols

    # ── 全量清洗 ──
    df_clean = clean_dataframe(df_data, patterns, dash_clean_cols)

    # ── Step 3: 输出 ──
    # 默认输出到桌面
    if not output_dir:
        output_dir = desktop or str(Path.home() / "Desktop")
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    clean_suffix = "_cleaned.csv"
    # CSV 输出必须带 BOM，否则 Excel 按本地编码（GBK）解析中文乱码
    csv_encoding = "utf-8-sig" if target_encoding.lower() in ("utf-8", "utf8") else target_encoding
    out_csv_path = make_output_path(file_path, clean_suffix, output_dir)
    write_csv(df_clean, out_csv_path, csv_encoding, text_format_cols)
    result["output_csv"] = out_csv_path

    if to_xlsx:
        xlsx_suffix = "_cleaned.xlsx"
        out_xlsx_path = make_output_path(file_path, xlsx_suffix, output_dir)
        write_xlsx(df_clean, out_xlsx_path, text_format_cols, date_format_cols, numeric_format_cols)
        result["output_xlsx"] = out_xlsx_path

    log.info(f"\n✅ 处理完成！")
    return result


def process_directory(
    dir_path: str,
    target_encoding: str = "utf-8-sig",
    to_xlsx: bool = False,
    output_dir: str = None,
    desktop: str = None,
) -> list:
    """批量处理目录下所有 .csv 文件。"""
    p = Path(dir_path)
    csv_files = sorted(p.glob("*.csv"))
    if not csv_files:
        log.warning(f"目录下没有找到 .csv 文件：{dir_path}")
        return []

    results = []
    for f in csv_files:
        try:
            r = process_csv(str(f), target_encoding, to_xlsx, output_dir, desktop=desktop)
            results.append(r)
        except Exception as e:
            log.error(f"  ❌ 处理失败：{e}")
            results.append({"file": str(f), "status": "error", "error": str(e)})

    return results


def print_summary(results: list | dict):
    """打印处理摘要。"""
    if isinstance(results, dict):
        results = [results]

    print(f"\n{'='*60}")
    print(f"📊 处理摘要")
    print(f"{'='*60}")
    ok = sum(1 for r in results if r.get("status") == "ok")
    err = sum(1 for r in results if r.get("status") == "error")
    print(f"  成功：{ok}  失败：{err}")
    for r in results:
        status = "✅" if r.get("status") == "ok" else "❌"
        fname = Path(r.get("file", "?")).name
        enc = r.get("encoding", "?")
        shape = r.get("shape", "?")
        txt_cols = r.get("text_format_cols", [])
        txt_info = f" 文本列：{txt_cols}" if txt_cols else ""
        print(f"  {status} {fname}  [{enc}]  {shape}  {txt_info}")
        if r.get("output_csv"):
            print(f"     CSV → {r['output_csv']}")
        if r.get("output_xlsx"):
            print(f"     XLSX → {r['output_xlsx']}")
        if r.get("error"):
            print(f"     Error: {r['error']}")


# ═══════════════════════════════════════════════════
#  CLI 入口
# ═══════════════════════════════════════════════════

def main():
    parser = argparse.ArgumentParser(
        description="qinshihuang-csv — 电商CSV大一统清洗方案"
    )
    parser.add_argument("input", nargs="?", help="CSV文件路径或目录路径")
    parser.add_argument("--encoding", default="utf-8-sig", help="目标编码 (默认 utf-8-sig)")
    parser.add_argument("--to-xlsx", action="store_true", help="额外输出 .xlsx 文件")
    parser.add_argument("--output-dir", help="输出目录（默认桌面）")
    parser.add_argument("--batch", action="store_true", help="批量模式：输入为目录时自动扫描所有 .csv")

    args = parser.parse_args()

    if not args.input:
        parser.print_help()
        return

    input_path = Path(args.input)
    if not input_path.exists():
        print(f"❌ 路径不存在：{args.input}")
        sys.exit(1)

    desktop = str(Path.home() / "Desktop")

    if input_path.is_dir():
        if not args.batch:
            print(f"📁 检测到目录，使用 --batch 批量处理所有 .csv 文件")
            print(f"   目录：{input_path}")
        results = process_directory(
            str(input_path), args.encoding, args.to_xlsx, args.output_dir, desktop
        )
    else:
        result = process_csv(
            str(input_path), args.encoding, args.to_xlsx, args.output_dir, desktop=desktop
        )
        results = [result]

    print_summary(results)


if __name__ == "__main__":
    main()
