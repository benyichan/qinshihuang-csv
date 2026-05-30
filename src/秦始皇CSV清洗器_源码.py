#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
qinshihuang-csv 图形化桌面版
============================
基于原清洗脚本封装，提供完整的 GUI 操作界面。
支持单个/批量 CSV 清洗，编码自动检测，脏数据清理，输出 CSV/XLSX。
"""

import os
import re
import sys
import codecs
import logging
import threading
import queue
import datetime
from pathlib import Path

# 必须在导入tkinter之前处理编码问题
import locale
import io
if sys.stdout is not None:
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding=locale.getpreferredencoding())
if sys.stderr is not None:
    sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding=locale.getpreferredencoding())

from tkinter import *
from tkinter import filedialog, messagebox, scrolledtext
from tkinter.ttk import Progressbar, Combobox, Checkbutton

# 核心依赖 - 必须确保这些被正确打包
import pandas as pd
from charset_normalizer import from_path as cn_from_path
import chardet
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

# 可选 clevercsv 不影响主体功能
try:
    import clevercsv
    HAS_CLEVERCSV = True
except ImportError:
    HAS_CLEVERCSV = False

# ──────────────────────────────────────────────
# 原始清洗脚本核心 (未作改动，仅整合)
# ──────────────────────────────────────────────
_CURRENCY_SYMBOLS = "￥$¥€£₩₽₪₫₱₹₴₸₺₼₿₠₡₢₣₤₥₦₧₨₰₲₳₵₶₷₻₾"
_STRIP_CHARS = f",;\t\n\r {_CURRENCY_SYMBOLS}"

FALLBACK_ENCODINGS = [
    "utf-8-sig", "utf-8", "gb18030", "gbk", "gb2312",
    "big5", "big5hkscs",
    "shift_jis", "euc-jp", "iso-2022-jp",
    "euc-kr", "iso-2022-kr",
    "utf-16", "utf-16-le", "utf-16-be",
    "utf-32", "utf-32-le", "utf-32-be",
    "cp1252", "cp1250", "cp1251", "cp1253", "cp1254", "cp1255", "cp1256", "cp1257", "cp1258",
    "iso-8859-1", "iso-8859-2", "iso-8859-3", "iso-8859-4", "iso-8859-5",
    "iso-8859-6", "iso-8859-7", "iso-8859-8", "iso-8859-9", "iso-8859-10",
    "iso-8859-13", "iso-8859-14", "iso-8859-15", "iso-8859-16",
    "latin1", "mac_roman", "mac_cyrillic", "mac_greek", "mac_iceland", "mac_turkish",
    "cp437", "cp850", "cp852", "cp855", "cp857", "cp858", "cp860", "cp861",
    "cp862", "cp863", "cp864", "cp865", "cp866", "cp869",
    "koi8-r", "koi8-u",
    "hz", "iso-2022-cn", "euc-tw",
]

def detect_bom(file_path: str):
    with open(file_path, "rb") as f:
        raw = f.read(4)
    if raw.startswith(codecs.BOM_UTF8):
        return "utf-8-sig"
    if raw[:4] == b"\xff\xfe\x00\x00":
        return "utf-32-le"
    if raw[:4] == b"\x00\x00\xfe\xff":
        return "utf-32-be"
    if raw[:2] == codecs.BOM_UTF16_LE:
        return "utf-16-le"
    if raw[:2] == codecs.BOM_UTF16_BE:
        return "utf-16-be"
    return None

def detect_encoding_charset_normalizer(file_path: str, sample_size: int = 100_000):
    try:
        with open(file_path, "rb") as f:
            raw_sample = f.read(sample_size)
        result = cn_from_path(file_path).best()
        if result is None:
            from charset_normalizer import from_bytes
            result = from_bytes(raw_sample).best()
        if result and result.encoding and result.confidence is not None:
            if result.confidence >= 0.7:
                enc = result.encoding.lower()
                if enc in ("gb2312", "gbk"):
                    return "gb18030"
                return enc
    except Exception:
        pass
    return None

def detect_encoding_chardet(file_path: str, sample_size: int = 100_000):
    try:
        with open(file_path, "rb") as f:
            raw = f.read(sample_size)
        result = chardet.detect(raw)
        enc = result.get("encoding")
        confidence = result.get("confidence", 0)
        lang = result.get("language", "")
        if enc:
            threshold = 0.2 if lang == "zh" else 0.5
            if confidence >= threshold:
                enc = enc.lower()
                if enc in ("gb2312", "gbk"):
                    return "gb18030"
                return enc
    except Exception:
        pass
    return None

def try_fallback_encodings(file_path: str):
    with open(file_path, "rb") as f:
        raw_header = f.read(4096)
    cjk_encodings = {"gb18030", "gbk", "gb2312", "big5", "big5hkscs", "euc-tw", "euc-jp", "shift_jis"}
    for enc in FALLBACK_ENCODINGS:
        try:
            raw_header.decode(enc)
            with open(file_path, "r", encoding=enc) as f:
                f.readline()
                f.readline()
            return enc
        except (UnicodeDecodeError, LookupError, UnicodeError):
            if enc in cjk_encodings:
                try:
                    decoded = raw_header.decode(enc, errors="replace")
                    cjk_count = sum(1 for c in decoded if '\u4e00' <= c <= '\u9fff')
                    if cjk_count >= 5:
                        return enc
                except Exception:
                    continue
    return None

def detect_encoding(file_path: str) -> str:
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"文件不存在：{file_path}")
    if path.stat().st_size == 0:
        raise ValueError(f"文件为空：{file_path}")
    bom_enc = detect_bom(file_path)
    if bom_enc:
        return bom_enc
    cn_enc = detect_encoding_charset_normalizer(file_path)
    if cn_enc:
        return cn_enc
    cd_enc = detect_encoding_chardet(file_path)
    if cd_enc:
        return cd_enc
    fb_enc = try_fallback_encodings(file_path)
    if fb_enc:
        return fb_enc
    raise ValueError(f"无法识别文件的编码格式：{file_path}")

def read_csv_with_encoding(file_path: str, encoding: str):
    sep = ","
    if HAS_CLEVERCSV:
        try:
            dialect = clevercsv.detect_dialect(file_path, encoding=encoding)
            sep = dialect.delimiter
        except Exception:
            pass
    if not sep or sep == "\x00":
        sep = ","
    with open(file_path, "r", encoding=encoding) as f:
        first_line = f.readline()
    if sep == "," and "\t" in first_line and "," not in first_line[:200]:
        sep = "\t"
    for errors_mode in ["strict", "replace"]:
        try:
            df = pd.read_csv(
                file_path,
                encoding=encoding,
                encoding_errors=errors_mode,
                sep=sep,
                dtype=str,
                keep_default_na=False,
                na_filter=False,
                header=None,
                on_bad_lines="skip",
            )
            break
        except UnicodeDecodeError:
            if errors_mode == "replace":
                continue
    return df, sep

def auto_detect_header(df: pd.DataFrame) -> int:
    for row_idx in range(min(5, len(df))):
        row = df.iloc[row_idx]
        all_clean = True
        for val in row:
            val_str = str(val).strip()
            if not val_str:
                continue
            if re.match(r"^[\d\s,;\t￥$¥€£]+$", val_str):
                all_clean = False
                break
        if all_clean and row.notna().sum() >= 2:
            return row_idx
    return 0

def clean_cell_text(text: str) -> str:
    if not isinstance(text, str):
        return str(text) if text is not None else ""
    if not text:
        return text
    text = text.replace("\t", "").replace("\n", "").replace("\r", "")
    prev = None
    while prev != text:
        prev = text
        text = text.strip(_STRIP_CHARS)
    return text

def is_pure_digits_15plus(s: str) -> bool:
    return bool(re.match(r"^\d{12,}$", s))

_DATE_PATTERNS = [
    r"^\d{4}-\d{1,2}-\d{1,2}$",
    r"^\d{4}/\d{1,2}/\d{1,2}$",
    r"^\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{2}$",
    r"^\d{4}-\d{1,2}-\d{1,2} \d{1,2}:\d{2}:\d{2}$",
    r"^\d{4}/\d{1,2}/\d{1,2} \d{1,2}:\d{2}:\d{2}$",
]

def is_date_string(s: str) -> bool:
    return any(re.match(p, s) for p in _DATE_PATTERNS)

def is_numeric_string(s: str) -> bool:
    if not s or s.strip() in ("", "-", "–"):
        return False
    try:
        float(s)
        return True
    except ValueError:
        return False

def is_pure_dash(s: str) -> bool:
    return s.strip() in ("-", "–", "—")

def scan_dirty_patterns(df: pd.DataFrame, sample_rows: int = 15):
    sample = df.head(sample_rows).copy()
    patterns = {}
    for col in df.columns:
        col_data = sample[col].dropna().astype(str)
        needs_clean = False
        text_format = False
        date_format = False
        numeric_format = False
        has_decimal = False
        has_dash = False
        for val in col_data:
            if not val:
                continue
            cleaned = clean_cell_text(val)
            if cleaned != val:
                needs_clean = True
            if is_pure_digits_15plus(cleaned):
                text_format = True
            if is_date_string(cleaned) or is_date_string(val):
                date_format = True
            if is_numeric_string(cleaned):
                numeric_format = True
            if "." in cleaned and is_numeric_string(cleaned):
                has_decimal = True
            if is_pure_dash(cleaned) or is_pure_dash(val):
                has_dash = True
        patterns[col] = {
            "needs_clean": needs_clean,
            "text_format": text_format,
            "date_format": date_format,
            "numeric_format": numeric_format,
            "has_decimal": has_decimal,
            "has_dash": has_dash,
        }
    text_format_cols = [c for c, p in patterns.items() if p["text_format"]]
    date_format_cols = [c for c, p in patterns.items() if p["date_format"] and not p["text_format"]]
    date_or_text_set = set(text_format_cols) | set(date_format_cols)
    numeric_format_cols = [c for c, p in patterns.items()
                           if p.get("has_decimal") and c not in date_or_text_set]
    dash_clean_cols = [c for c, p in patterns.items()
                       if p.get("has_dash") and c not in date_or_text_set]
    return patterns, text_format_cols, date_format_cols, numeric_format_cols, dash_clean_cols

def clean_dataframe(df: pd.DataFrame, patterns: dict, dash_clean_cols: list) -> pd.DataFrame:
    df_clean = df.copy()
    total_cells = 0
    cleaned_cells = 0
    dash_fixes = 0
    dash_set = set(dash_clean_cols)
    for col in df_clean.columns:
        for idx in range(len(df_clean)):
            orig = df_clean.at[idx, col]
            if not isinstance(orig, str) or not orig:
                continue
            total_cells += 1
            cleaned = clean_cell_text(orig)
            if cleaned != orig:
                df_clean.at[idx, col] = cleaned
                cleaned_cells += 1
            if col in dash_set and is_pure_dash(cleaned):
                df_clean.at[idx, col] = "0"
                dash_fixes += 1
    return df_clean

def make_output_path(original_path: str, suffix: str, output_dir: str = None) -> str:
    p = Path(original_path)
    stem = p.stem
    if output_dir:
        out_dir = Path(output_dir)
    else:
        out_dir = p.parent
    out_dir.mkdir(parents=True, exist_ok=True)
    return str(out_dir / f"{stem}{suffix}")

def write_csv(df: pd.DataFrame, output_path: str, encoding: str = "utf-8-sig", text_format_cols: list = None):
    df_out = df.copy()
    text_set = set(text_format_cols or [])
    if text_set:
        for col in df_out.columns:
            if col in text_set:
                df_out[col] = df_out[col].apply(lambda v: f"'{v}" if v and str(v).strip() else v)
    df_out.to_csv(output_path, index=False, encoding=encoding)

def _parse_date_safe(s):
    if not s or not isinstance(s, str):
        return s
    s = s.strip()
    for fmt in [
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d",
    ]:
        try:
            return datetime.datetime.strptime(s, fmt)
        except ValueError:
            continue
    return s

def write_xlsx(df: pd.DataFrame, output_path: str, text_format_cols: list = None,
               date_format_cols: list = None, numeric_format_cols: list = None):
    """写入Excel文件，openpyxl已在模块级别导入"""
    wb = Workbook()
    ws = wb.active
    headers = list(df.columns)
    ws.append(headers)
    text_set = set(text_format_cols or [])
    date_set = set(date_format_cols or [])
    numeric_set = set(numeric_format_cols or [])
    text_col_indices = {i for i, h in enumerate(headers, 1) if h in text_set}
    date_col_indices = {i for i, h in enumerate(headers, 1) if h in date_set and h not in text_set}
    numeric_col_indices = {i for i, h in enumerate(headers, 1)
                           if h in numeric_set and h not in text_set and h not in date_set}
    for _, row in df.iterrows():
        values = []
        for col_idx, (col_name, val) in enumerate(zip(headers, row), start=1):
            if col_idx in date_col_indices:
                values.append(_parse_date_safe(val))
            elif col_idx in numeric_col_indices:
                try:
                    v = float(val) if val and str(val).strip() not in ("", "-", "–") else 0.0
                    values.append(v)
                except (ValueError, TypeError):
                    values.append(0.0)
            elif col_idx in text_col_indices:
                values.append(val)
            else:
                values.append(val)
        ws.append(values)
    for col_idx in text_col_indices:
        col_letter = get_column_letter(col_idx)
        for row_num in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row_num}"].number_format = '@'
    for col_idx in date_col_indices:
        col_letter = get_column_letter(col_idx)
        for row_num in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row_num}"].number_format = 'YYYY-MM-DD HH:MM:SS'
    for col_idx in numeric_col_indices:
        col_letter = get_column_letter(col_idx)
        for row_num in range(2, ws.max_row + 1):
            ws[f"{col_letter}{row_num}"].number_format = '#,##0.00'
    for col_idx, col_name in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        max_len = len(str(col_name))
        for row_num in range(2, min(ws.max_row + 1, 102)):
            cell_val = ws[f"{col_letter}{row_num}"].value
            if cell_val:
                max_len = max(max_len, len(str(cell_val)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 50)
    wb.save(output_path)

# 原始单文件处理函数（增加日志回调参数，保持兼容）
def process_csv(
    file_path: str,
    target_encoding: str = "utf-8-sig",
    to_xlsx: bool = False,
    output_dir: str = None,
    log_callback=None,
) -> dict:
    """处理单个 CSV 文件，log_callback 可选用于 GUI 日志"""
    def log_msg(msg, level="info"):
        if log_callback:
            log_callback(msg, level)
        else:
            print(msg)
    log_msg(f"\n{'='*60}")
    log_msg(f"📄 处理文件：{file_path}")
    log_msg(f"{'='*60}")
    result = {
        "file": file_path,
        "status": "ok",
        "encoding": None,
        "shape": None,
        "cleaned_cells": 0,
        "text_format_cols": [],
        "output_csv": None,
        "output_xlsx": None,
    }
    try:
        detected_enc = detect_encoding(file_path)
        result["encoding"] = detected_enc
        log_msg(f"🎯 最终编码：{detected_enc}")
        df, sep = read_csv_with_encoding(file_path, detected_enc)
        result["shape"] = f"{df.shape[0]}行×{df.shape[1]}列"
        if df.empty:
            log_msg("⚠️ 文件为空，跳过清洗")
            result["status"] = "empty"
            return result
        # 自动识别表头
        header_row = auto_detect_header(df)
        log_msg(f"📋 表头所在行：第 {header_row + 1} 行")
        headers = [str(v).strip() for v in df.iloc[header_row].tolist()]
        df.columns = headers
        df_data = df.iloc[header_row + 1:].reset_index(drop=True)
        patterns, text_format_cols, date_format_cols, numeric_format_cols, dash_clean_cols = scan_dirty_patterns(df_data)
        result["text_format_cols"] = text_format_cols
        df_clean = clean_dataframe(df_data, patterns, dash_clean_cols)
        if not output_dir:
            output_dir = str(Path.home() / "Desktop")
        Path(output_dir).mkdir(parents=True, exist_ok=True)
        csv_encoding = "utf-8-sig" if target_encoding.lower() in ("utf-8", "utf8") else target_encoding
        out_csv_path = make_output_path(file_path, "_cleaned.csv", output_dir)
        write_csv(df_clean, out_csv_path, csv_encoding, text_format_cols)
        result["output_csv"] = out_csv_path
        log_msg(f"✅ CSV 输出：{out_csv_path}")
        if to_xlsx:
            out_xlsx_path = make_output_path(file_path, "_cleaned.xlsx", output_dir)
            write_xlsx(df_clean, out_xlsx_path, text_format_cols, date_format_cols, numeric_format_cols)
            result["output_xlsx"] = out_xlsx_path
            log_msg(f"✅ XLSX 输出：{out_xlsx_path}")
        log_msg(f"🎉 文件处理完成: {file_path}")
    except Exception as e:
        result["status"] = "error"
        result["error"] = str(e)
        log_msg(f"❌ 处理失败：{e}", "error")
    return result


# ──────────────────────────────────────────────
# GUI 应用程序
# ──────────────────────────────────────────────
class CleaningApp:
    def __init__(self, root):
        self.root = root
        self.root.title("秦始皇 CSV 清洗器 - 电商数据大一统")
        self.root.geometry("1000x700")
        self.center_window()
        self.file_list = []   # 待处理文件路径列表
        self.stop_flag = False
        self.processing_thread = None

        # 设置样式
        self.create_widgets()
        # 重定向日志
        self.setup_logging()

    def center_window(self):
        self.root.update_idletasks()
        w = 1000
        h = 700
        x = (self.root.winfo_screenwidth() // 2) - (w // 2)
        y = (self.root.winfo_screenheight() // 2) - (h // 2)
        self.root.geometry(f"{w}x{h}+{x}+{y}")

    def create_widgets(self):
        # 主框架
        main_frame = Frame(self.root, padx=10, pady=10)
        main_frame.pack(fill=BOTH, expand=True)

        # 文件选择区域
        file_frame = LabelFrame(main_frame, text="待处理文件", padx=5, pady=5)
        file_frame.pack(fill=BOTH, expand=True, pady=(0,10))

        # 按钮栏
        btn_frame = Frame(file_frame)
        btn_frame.pack(fill=X, pady=5)
        Button(btn_frame, text="[+] 添加CSV文件", command=self.add_files, width=15).pack(side=LEFT, padx=2)
        Button(btn_frame, text="[D] 添加文件夹", command=self.add_folder, width=15).pack(side=LEFT, padx=2)
        Button(btn_frame, text="[-] 移除选中", command=self.remove_selected, width=12).pack(side=LEFT, padx=2)
        Button(btn_frame, text="[X] 清空列表", command=self.clear_list, width=10).pack(side=LEFT, padx=2)

        # 文件列表框 + 滚动条
        list_frame = Frame(file_frame)
        list_frame.pack(fill=BOTH, expand=True)
        scrollbar = Scrollbar(list_frame)
        scrollbar.pack(side=RIGHT, fill=Y)
        self.file_listbox = Listbox(list_frame, selectmode=EXTENDED, yscrollcommand=scrollbar.set)
        self.file_listbox.pack(side=LEFT, fill=BOTH, expand=True)
        scrollbar.config(command=self.file_listbox.yview)

        # 输出设置区域
        output_frame = LabelFrame(main_frame, text="输出设置", padx=5, pady=5)
        output_frame.pack(fill=X, pady=(0,10))

        # 输出目录选择
        out_dir_frame = Frame(output_frame)
        out_dir_frame.pack(fill=X, pady=3)
        Label(out_dir_frame, text="输出目录:").pack(side=LEFT)
        self.output_dir_var = StringVar(value=str(Path.home() / "Desktop"))
        Entry(out_dir_frame, textvariable=self.output_dir_var, width=50).pack(side=LEFT, padx=5, fill=X, expand=True)
        Button(out_dir_frame, text="浏览...", command=self.select_output_dir).pack(side=LEFT, padx=2)

        # 格式选项
        opt_frame = Frame(output_frame)
        opt_frame.pack(fill=X, pady=5)
        Label(opt_frame, text="目标编码:").pack(side=LEFT)
        self.encoding_var = StringVar(value="utf-8-sig")
        enc_combo = Combobox(opt_frame, textvariable=self.encoding_var, values=["utf-8-sig", "utf-8", "gb18030", "gbk"], width=12)
        enc_combo.pack(side=LEFT, padx=10)
        self.to_xlsx_var = BooleanVar(value=False)
        Checkbutton(opt_frame, text="同时输出 XLSX 文件", variable=self.to_xlsx_var).pack(side=LEFT, padx=10)

        # 大文件阈值
        thresh_frame = Frame(output_frame)
        thresh_frame.pack(fill=X, pady=3)
        Label(thresh_frame, text="大文件警告阈值 (MB):").pack(side=LEFT)
        self.threshold_var = StringVar(value="200")
        Entry(thresh_frame, textvariable=self.threshold_var, width=8).pack(side=LEFT, padx=5)
        Label(thresh_frame, text="超过此大小会弹出确认对话框", fg="gray").pack(side=LEFT, padx=5)

        # 进度条及开始按钮
        action_frame = Frame(main_frame)
        action_frame.pack(fill=X, pady=5)
        self.progress = Progressbar(action_frame, mode='indeterminate', length=300)
        self.progress.pack(side=LEFT, padx=5, fill=X, expand=True)
        self.start_btn = Button(action_frame, text="▶ 开始清洗", command=self.start_processing, bg="#4caf50", fg="white", font=("",10,"bold"))
        self.start_btn.pack(side=RIGHT, padx=5)
        self.stop_btn = Button(action_frame, text="■ 停止", command=self.stop_processing, state=DISABLED, bg="#f44336", fg="white")
        self.stop_btn.pack(side=RIGHT, padx=5)

        # 日志显示区域
        log_frame = LabelFrame(main_frame, text="运行日志", padx=5, pady=5)
        log_frame.pack(fill=BOTH, expand=True)
        self.log_text = scrolledtext.ScrolledText(log_frame, height=15, wrap=WORD, font=("Consolas", 9))
        self.log_text.pack(fill=BOTH, expand=True)
        # 清空日志按钮
        btn_clear_log = Button(log_frame, text="清空日志", command=self.clear_log)
        btn_clear_log.pack(pady=2)

    def setup_logging(self):
        """配置 logging 重定向到 GUI 文本控件"""
        self.log_queue = queue.Queue()
        self.root.after(100, self.poll_log_queue)
        # 自定义 Handler
        class QueueHandler(logging.Handler):
            def __init__(self, log_queue):
                super().__init__()
                self.log_queue = log_queue
            def emit(self, record):
                self.log_queue.insert(0, self.format(record))
        # 获取根日志记录器，避免重复添加
        logger = logging.getLogger()
        for hdlr in logger.handlers[:]:
            if isinstance(hdlr, QueueHandler):
                logger.removeHandler(hdlr)
        qh = QueueHandler(self.log_queue)
        formatter = logging.Formatter("%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
        qh.setFormatter(formatter)
        logger.addHandler(qh)
        logger.setLevel(logging.INFO)
        # 也保留控制台输出便于调试
        self.log_callback = self.append_log

    def append_log(self, msg, level="info"):
        """供 process_csv 调用的日志回调"""
        # 级别映射
        if level == "error":
            tag = "error"
        else:
            tag = "info"
        self.log_queue.put(msg)

    def poll_log_queue(self):
        """定期从队列取出日志并显示"""
        while not self.log_queue.empty():
            msg = self.log_queue.get()
            self.log_text.insert(END, msg + "\n")
            self.log_text.see(END)
        self.root.after(200, self.poll_log_queue)

    def clear_log(self):
        self.log_text.delete(1.0, END)

    def add_files(self):
        files = filedialog.askopenfilenames(
            title="选择CSV文件",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")]
        )
        for f in files:
            if f not in self.file_list:
                self.file_list.append(f)
                self.file_listbox.insert(END, f)
        if len(files):
            self.append_log(f"已添加 {len(files)} 个文件", "info")

    def add_folder(self):
        folder = filedialog.askdirectory(title="选择包含CSV文件的文件夹")
        if not folder:
            return
        # 是否包含子文件夹？增加简单选项，默认递归
        recursive = messagebox.askyesno("递归扫描", "是否包含子文件夹中的CSV文件？", parent=self.root)
        pattern = "**/*.csv" if recursive else "*.csv"
        csv_files = list(Path(folder).glob(pattern))
        added = 0
        for f in csv_files:
            fstr = str(f)
            if fstr not in self.file_list:
                self.file_list.append(fstr)
                self.file_listbox.insert(END, fstr)
                added += 1
        self.append_log(f"从文件夹添加 {added} 个CSV文件 (递归={recursive})", "info")

    def remove_selected(self):
        selected = self.file_listbox.curselection()
        for idx in reversed(selected):
            del self.file_list[idx]
            self.file_listbox.delete(idx)
        self.append_log(f"移除 {len(selected)} 个文件", "info")

    def clear_list(self):
        self.file_list.clear()
        self.file_listbox.delete(0, END)
        self.append_log("文件列表已清空", "info")

    def select_output_dir(self):
        dir_path = filedialog.askdirectory(title="选择输出目录")
        if dir_path:
            self.output_dir_var.set(dir_path)

    def check_large_file(self, filepath, threshold_mb):
        size_mb = os.path.getsize(filepath) / (1024 * 1024)
        if size_mb > threshold_mb:
            ret = messagebox.askyesno(
                "大文件警告",
                f"文件 {Path(filepath).name} 大小为 {size_mb:.1f} MB，超过阈值 {threshold_mb} MB。\n处理可能导致运行缓慢或内存不足，是否继续？",
                parent=self.root
            )
            return ret
        return True

    def start_processing(self):
        if not self.file_list:
            messagebox.showwarning("无文件", "请先添加待处理的 CSV 文件。")
            return
        # 获取阈值
        try:
            threshold = float(self.threshold_var.get())
        except ValueError:
            threshold = 200
        # 预检大文件
        for fp in self.file_list:
            if not self.check_large_file(fp, threshold):
                return

        self.stop_flag = False
        self.start_btn.config(state=DISABLED)
        self.stop_btn.config(state=NORMAL)
        self.progress.start()
        self.append_log("🚀 开始批量清洗任务...", "info")
        # 启动后台线程
        self.processing_thread = threading.Thread(target=self.process_all_files, daemon=True)
        self.processing_thread.start()

    def stop_processing(self):
        self.stop_flag = True
        self.append_log("⚠️ 用户请求停止，将在当前文件完成后结束...", "warning")

    def process_all_files(self):
        output_dir = self.output_dir_var.get().strip()
        if not output_dir:
            output_dir = str(Path.home() / "Desktop")
        to_xlsx = self.to_xlsx_var.get()
        target_enc = self.encoding_var.get()
        success = 0
        error = 0
        for idx, filepath in enumerate(self.file_list):
            if self.stop_flag:
                self.append_log("⏹️ 任务已手动终止", "warning")
                break
            self.append_log(f"\n【{idx+1}/{len(self.file_list)}】开始处理: {Path(filepath).name}")
            # 调用清洗函数，传入日志回调
            res = process_csv(
                file_path=filepath,
                target_encoding=target_enc,
                to_xlsx=to_xlsx,
                output_dir=output_dir,
                log_callback=self.append_log
            )
            if res.get("status") == "ok":
                success += 1
            else:
                error += 1
        self.root.after(0, self.finish_processing, success, error)

    def finish_processing(self, success, error):
        self.progress.stop()
        self.start_btn.config(state=NORMAL)
        self.stop_btn.config(state=DISABLED)
        self.append_log(f"\n✅ 批量处理完成！成功: {success}  失败: {error}", "info")
        messagebox.showinfo("完成", f"清洗任务结束\n成功: {success}\n失败: {error}", parent=self.root)

def main():
    root = Tk()
    app = CleaningApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()